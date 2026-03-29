#!/usr/bin/env python3
"""
Build TikTok account credentials from an input XLSX file.

Input:  an XLSX sheet with at least `phone` and `api_phone` columns.
Output: same sheet data, plus ensured columns:
        - phone
        - api_phone
        - username
        - password
        - cookie_tiktok

Notes:
- Username/password are derived from phone + api_phone and are deterministic.
- Existing username/password/cookie_tiktok values are kept by default.
"""

from __future__ import annotations

import argparse
import hashlib
import re
import sys
from pathlib import Path
from typing import Dict, Tuple


HEADER_ALIASES = {
    "phone": {"phone", "phone_number", "so_dien_thoai", "sdt"},
    "api_phone": {"api_phone", "api-phone", "api phone", "phone_api"},
    "username": {"username", "user_name", "user"},
    "password": {"password", "pass", "pwd"},
    "cookie_tiktok": {"cookie_tiktok", "cookie tiktok", "tiktok_cookie", "cookie"},
}

SPECIAL_CHARS = "!@#$%&*?"
ALPHA = "abcdefghijklmnopqrstuvwxyz"


def normalize_header(value: object) -> str:
    raw = "" if value is None else str(value)
    raw = raw.strip().lower()
    raw = raw.replace("-", "_")
    raw = re.sub(r"\s+", "_", raw)
    return raw


def digits_only(value: object) -> str:
    return re.sub(r"\D+", "", "" if value is None else str(value))


def normalize_us_local(phone: object) -> str:
    digits = digits_only(phone)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return digits


def to_us_e164(local_digits: str) -> str:
    if not local_digits:
        return ""
    if len(local_digits) >= 10:
        local_digits = local_digits[-10:]
    return f"+1{local_digits}"


def pick_from_digest(digest: str, start: int, modulus: int) -> int:
    return int(digest[start : start + 2], 16) % modulus


def make_username(
    local_digits: str,
    api_phone: str,
    seed: str,
    row_number: int,
    used: Dict[str, int],
) -> str:
    core_digits = local_digits[-6:] if local_digits else str(row_number).zfill(6)
    digest = hashlib.sha256(
        f"{local_digits}|{api_phone}|{seed}|username".encode("utf-8")
    ).hexdigest()
    suffix = (
        ALPHA[pick_from_digest(digest, 0, len(ALPHA))]
        + ALPHA[pick_from_digest(digest, 2, len(ALPHA))]
    )
    base = f"tt{core_digits}{suffix}"
    if base not in used:
        used[base] = 1
        return base
    used[base] += 1
    return f"{base}{used[base]}"


def make_password(local_digits: str, api_phone: str, seed: str) -> str:
    last4 = (local_digits[-4:] if local_digits else "0000").rjust(4, "0")
    digest = hashlib.sha256(
        f"{local_digits}|{api_phone}|{seed}|password".encode("utf-8")
    ).hexdigest()
    upper = ALPHA[pick_from_digest(digest, 0, len(ALPHA))].upper()
    lower1 = ALPHA[pick_from_digest(digest, 2, len(ALPHA))]
    lower2 = ALPHA[pick_from_digest(digest, 4, len(ALPHA))]
    digit1 = str(pick_from_digest(digest, 6, 10))
    digit2 = str(pick_from_digest(digest, 8, 10))
    symbol2 = SPECIAL_CHARS[pick_from_digest(digest, 12, len(SPECIAL_CHARS))]
    tail = ALPHA[pick_from_digest(digest, 14, len(ALPHA))].upper()
    # Contains uppercase, lowercase, numbers, and special chars.
    return f"{upper}{lower1}!{last4}{lower2}{digit1}{symbol2}{digit2}{tail}"


def find_header_columns(sheet) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        header_key = normalize_header(sheet.cell(row=1, column=col).value)
        if not header_key:
            continue
        for canonical, aliases in HEADER_ALIASES.items():
            if header_key in aliases:
                mapping.setdefault(canonical, col)
    return mapping


def ensure_column(sheet, mapping: Dict[str, int], canonical_name: str) -> int:
    existing = mapping.get(canonical_name)
    if existing:
        return existing
    col = sheet.max_column + 1
    sheet.cell(row=1, column=col).value = canonical_name
    mapping[canonical_name] = col
    return col


def process_sheet(
    sheet,
    seed: str,
    overwrite_existing: bool,
) -> Tuple[int, int]:
    mapping = find_header_columns(sheet)

    if "phone" not in mapping:
        raise ValueError("Missing required column: phone")
    if "api_phone" not in mapping:
        raise ValueError("Missing required column: api_phone")

    phone_col = ensure_column(sheet, mapping, "phone")
    api_phone_col = ensure_column(sheet, mapping, "api_phone")
    username_col = ensure_column(sheet, mapping, "username")
    password_col = ensure_column(sheet, mapping, "password")
    cookie_col = ensure_column(sheet, mapping, "cookie_tiktok")

    used_usernames: Dict[str, int] = {}
    updated_rows = 0
    skipped_rows = 0

    for row in range(2, sheet.max_row + 1):
        raw_phone = sheet.cell(row=row, column=phone_col).value
        raw_api_phone = sheet.cell(row=row, column=api_phone_col).value

        local_digits = normalize_us_local(raw_phone)
        api_digits = normalize_us_local(raw_api_phone)

        if not local_digits and not api_digits:
            skipped_rows += 1
            continue

        if not local_digits and api_digits:
            local_digits = api_digits

        api_phone_value = (
            str(raw_api_phone).strip() if raw_api_phone is not None else ""
        )
        if not api_phone_value:
            api_phone_value = to_us_e164(local_digits)
            sheet.cell(row=row, column=api_phone_col).value = api_phone_value

        username_cell = sheet.cell(row=row, column=username_col)
        password_cell = sheet.cell(row=row, column=password_col)
        cookie_cell = sheet.cell(row=row, column=cookie_col)

        has_username = bool(str(username_cell.value).strip()) if username_cell.value else False
        has_password = bool(str(password_cell.value).strip()) if password_cell.value else False

        if overwrite_existing or not has_username:
            username_cell.value = make_username(
                local_digits=local_digits,
                api_phone=api_phone_value,
                seed=seed,
                row_number=row,
                used=used_usernames,
            )

        if overwrite_existing or not has_password:
            password_cell.value = make_password(
                local_digits=local_digits,
                api_phone=api_phone_value,
                seed=seed,
            )

        if cookie_cell.value is None:
            cookie_cell.value = ""

        updated_rows += 1

    return updated_rows, skipped_rows


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Read XLSX with phone/api_phone and output XLSX with "
            "username/password/cookie_tiktok."
        )
    )
    parser.add_argument("--input", required=True, help="Input .xlsx file path")
    parser.add_argument(
        "--output",
        help="Output .xlsx file path (default: <input>_with_accounts.xlsx)",
    )
    parser.add_argument(
        "--sheet",
        help="Sheet name (default: active sheet)",
    )
    parser.add_argument(
        "--seed",
        default="buglogin",
        help="Seed for deterministic username/password generation",
    )
    parser.add_argument(
        "--overwrite-existing",
        action="store_true",
        help="Overwrite existing username/password cells",
    )
    return parser


def ensure_openpyxl():
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError:
        print(
            "Missing dependency: openpyxl\n"
            "Install it with:\n"
            "  py -m pip install openpyxl\n"
            "or:\n"
            "  python -m pip install openpyxl",
            file=sys.stderr,
        )
        raise SystemExit(1)
    return load_workbook


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 1
    if input_path.suffix.lower() != ".xlsx":
        print("Only .xlsx is supported.", file=sys.stderr)
        return 1

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        output_path = input_path.with_name(f"{input_path.stem}_with_accounts.xlsx")

    load_workbook = ensure_openpyxl()
    workbook = load_workbook(filename=str(input_path))

    if args.sheet:
        if args.sheet not in workbook.sheetnames:
            print(
                f"Sheet '{args.sheet}' not found. Available: {', '.join(workbook.sheetnames)}",
                file=sys.stderr,
            )
            return 1
        sheet = workbook[args.sheet]
    else:
        sheet = workbook.active

    try:
        updated_rows, skipped_rows = process_sheet(
            sheet=sheet,
            seed=args.seed,
            overwrite_existing=args.overwrite_existing,
        )
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")
    print(f"Updated rows: {updated_rows}")
    print(f"Skipped rows (empty phone/api_phone): {skipped_rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
